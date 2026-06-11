"""Export service — generates CSV, Excel, and PDF reports for Telegram delivery."""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from pathlib import Path
from typing import Any, Sequence

from sqlalchemy.ext.asyncio import AsyncSession

from src.core.config.settings import get_settings
from src.core.services.analytics_service import AnalyticsService
from src.core.repositories.job_repository import JobRepository
from src.core.repositories.application_repository import ApplicationRepository
from src.core.repositories.company_repository import CompanyRepository


class ExportService:
    """Generates export reports in multiple formats."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.analytics = AnalyticsService(session)
        self.job_repo = JobRepository(session)
        self.app_repo = ApplicationRepository(session)
        self.company_repo = CompanyRepository(session)
        self.export_dir = get_settings().storage.export_dir

    async def export_csv(self, report_type: str = "all") -> Path:
        """Generate a CSV report.

        Args:
            report_type: 'jobs', 'applications', 'companies', 'salary', or 'all'
        """
        self.export_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.export_dir / f"report_{report_type}_{timestamp}.csv"

        output = io.StringIO()
        writer = csv.writer(output)

        if report_type in ("jobs", "all"):
            writer.writerow([
                "Job ID", "Title", "Company", "Location", "Provider",
                "Salary Est.", "Match Score", "Status", "Discovered"
            ])
            jobs = await self.job_repo.get_all(limit=1000)
            for job in jobs:
                writer.writerow([
                    str(job.id)[:8],
                    job.title,
                    "",  # Company name would need join
                    job.location or "N/A",
                    job.ats_provider,
                    f"₹{job.salary_estimate or 0}L",
                    f"{job.match_score or 0:.0f}%",
                    job.status.value,
                    job.discovered_at.strftime("%Y-%m-%d %H:%M") if job.discovered_at else "",
                ])

        if report_type in ("applications", "all"):
            if report_type == "all":
                writer.writerow([])  # separator
            writer.writerow([
                "App ID", "Job ID", "Status", "Method",
                "Match Score", "Salary Est.", "Applied At"
            ])
            apps = await self.app_repo.get_all(limit=1000)
            for app in apps:
                writer.writerow([
                    str(app.id)[:8],
                    str(app.job_id)[:8],
                    app.status.value,
                    app.method.value,
                    f"{app.match_score or 0:.0f}%",
                    f"₹{app.salary_estimate or 0}L",
                    app.applied_at.strftime("%Y-%m-%d %H:%M") if app.applied_at else "pending",
                ])

        filepath.write_text(output.getvalue(), encoding="utf-8")
        return filepath

    async def export_excel(self, report_type: str = "all") -> Path:
        """Generate an Excel report with multiple sheets."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel export")

        self.export_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.export_dir / f"report_{report_type}_{timestamp}.xlsx"

        wb = Workbook()

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")

        def style_header(ws: Any, headers: list[str]) -> None:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # Jobs sheet
        if report_type in ("jobs", "all"):
            ws_jobs = wb.active
            if ws_jobs is not None:
                ws_jobs.title = "Jobs"
                style_header(ws_jobs, [
                    "Title", "Location", "Provider", "Salary (LPA)",
                    "Match %", "Status", "Discovered"
                ])
                jobs = await self.job_repo.get_all(limit=1000)
                for job in jobs:
                    ws_jobs.append([
                        job.title,
                        job.location or "N/A",
                        job.ats_provider,
                        job.salary_estimate or 0,
                        job.match_score or 0,
                        job.status.value,
                        job.discovered_at.strftime("%Y-%m-%d") if job.discovered_at else "",
                    ])

        # Applications sheet
        if report_type in ("applications", "all"):
            ws_apps = wb.create_sheet("Applications")
            style_header(ws_apps, [
                "Status", "Method", "Match %", "Salary (LPA)", "Applied At"
            ])
            apps = await self.app_repo.get_all(limit=1000)
            for app in apps:
                ws_apps.append([
                    app.status.value,
                    app.method.value,
                    app.match_score or 0,
                    app.salary_estimate or 0,
                    app.applied_at.strftime("%Y-%m-%d %H:%M") if app.applied_at else "pending",
                ])

        # Stats sheet
        if report_type == "all":
            ws_stats = wb.create_sheet("Statistics")
            stats = await self.analytics.get_daily_stats()
            style_header(ws_stats, ["Metric", "Value"])
            for key, value in stats.items():
                ws_stats.append([key.replace("_", " ").title(), str(value)])

        wb.save(str(filepath))
        return filepath

    async def export_pdf(self, report_type: str = "all") -> Path:
        """Generate a PDF summary report."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            raise RuntimeError("reportlab is required for PDF export")

        self.export_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.export_dir / f"report_{report_type}_{timestamp}.pdf"

        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph("JobHunter AI — Report", styles["Title"]))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Normal"]
        ))
        elements.append(Spacer(1, 20))

        # Statistics
        stats = await self.analytics.get_daily_stats()
        stats_data = [["Metric", "Value"]]
        for key, value in stats.items():
            if key != "date":
                stats_data.append([key.replace("_", " ").title(), str(value)])

        stats_table = Table(stats_data, colWidths=[250, 200])
        stats_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        elements.append(Paragraph("Daily Statistics", styles["Heading2"]))
        elements.append(stats_table)
        elements.append(Spacer(1, 20))

        # Salary stats
        salary_stats = await self.analytics.get_salary_stats()
        elements.append(Paragraph("Salary Intelligence", styles["Heading2"]))
        salary_data = [["Metric", "Value"]]
        salary_data.append(["Average Salary", f"₹{salary_stats['average']}L"])
        salary_data.append(["Highest Salary", f"₹{salary_stats['highest']}L"])
        salary_data.append(["Lowest Salary", f"₹{salary_stats['lowest']}L"])
        salary_data.append(["Total Estimates", str(salary_stats['total_estimates'])])

        salary_table = Table(salary_data, colWidths=[250, 200])
        salary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#A23B72")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        elements.append(salary_table)

        doc.build(elements)
        return filepath
